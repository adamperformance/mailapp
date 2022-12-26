from datetime import datetime
from tkinter import *
from tkinter import ttk
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

import os

wb = load_workbook("GES.xlsx")
sheets = wb.sheetnames
comp_temp = wb[sheets[0]]
ca = wb[sheets[1]]
email = wb[sheets[2]]

templ_title = []
templ_subj = []
templ_bodytext = []

for i in range(2,5):
    templ_title.append(comp_temp[f"A{i}"].value)
    templ_subj.append(comp_temp[f"B{i}"].value)
    templ_bodytext.append(comp_temp[f"C{i}"].value)

# print(templ_title)   
# print(templ_subj)
# print(templ_bodytext)

manager = []
reviewer = []
preparer = []

for i in range (2,5):
    manager.append(ca[f"B{i}"].value)
    reviewer.append(ca[f"C{i}"].value)
    preparer.append(ca[f"D{i}"].value)

print(manager)
print(reviewer)
print(preparer)

body = templ_bodytext[1]
recipient = "whatever"
cc = f"{manager[0]}, {reviewer[0]}, {preparer[0]}"
subject = templ_subj[1]

print(body)

os.system("thunderbird -compose to=" + recipient + ",cc=" + cc + ",subject=" + subject + ",body=" + body)



# to create anything in tkinter, first you define then put it on the screen
# root = Tk()
# root.title("Étrend számoló")
# frame1 = Frame(root, relief=RIDGE, borderwidth=5)
# frame2 = Frame(root, relief=RIDGE, borderwidth=5)
# frame3 = Frame(root, relief=RIDGE, borderwidth=5)
# frame4 = Frame(root, relief=RIDGE, borderwidth=5)    


# BMR_lbl = Label(master=frame1, text = "Whatever")
# BMR_lbl.pack()
# button1 = Button(master=frame2,text="Megadom!")
# button1.pack()
# lbl = Label(master=frame3, text = "Label 3")
# lbl.pack()
# xxx = Button(master=frame4, text = "another text")
# xxx.pack()

# frame1.pack(fill=X)
# frame2.pack(fill=X)
# frame3.pack(fill=X)
# frame4.pack(fill=X)

# root.mainloop()