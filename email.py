# import win32com.client
from tkinter import *
from tkinter import ttk
from docx import Document
from openpyxl.workbook import Workbook
from openpyxl import load_workbook


# The layout of the code i the following:
    # Section 1: 
        # base python functionality of the code - whatever was needed to preprocess
        # the data (e.g. word & excel documents)
    # Section 2:
        # function definition for 
    # Section 3:
        # the tkinter code that creates the GUI layout of the program
    # Section 4:
        # email functionality



# SECTION 1

document = Document("ctp1.docx")

# get all the titles, subtitles and body texts.
templates = []
subjects = []   
curr_paragraph = ""
bodytexts = []
x = 0

for paragraph in document.paragraphs:
    if paragraph.style.name == "Heading 1":
        templates.append(paragraph.text)
        x = 1
    elif paragraph.style.name == "Heading 2":
        subjects.append(paragraph.text)
        x = 2
    elif paragraph.style.name == "Normal":
        x = 0

    if x == 0:
        curr_paragraph = curr_paragraph + paragraph.text
    elif x == 1:
        bodytexts.append(curr_paragraph)
        curr_paragraph = ""

# get the teams from the excel

wb = load_workbook("GES.xlsx")
sheets = wb.sheetnames
ca = wb[sheets[1]]
email = wb[sheets[2]]

eng_name = []
engagement = []
manager = []
reviewer = []
preparer = []
emails = []

for i in range (2,5):
    eng_name.append(ca[f"A{i}"].value)
    manager.append(ca[f"B{i}"].value)
    reviewer.append(ca[f"C{i}"].value)
    preparer.append(ca[f"D{i}"].value)
    # should turn this into a list of dictionaries where eng name and team and the team
    # is another dict for Man/Reviewer/Preparer



for i in range (2,5):
    name = email[f"A{i}"].value
    address = email[f"B{i}"].value
    emails.append({"name":name,"email":address})

def getTemplate():
    chosen_template = template_list.get()    
    temp_num = template_list.current()

    return chosen_template, temp_num

def getEngagement():
    chosen_engagement = engagement_list.get()
    eng_num = engagement_list.current()

    return chosen_engagement, eng_num

def get_emails():
    chosen_engagement = engagement_list.get()

    pass

def getRemaining():
    temp_chosen, temp_num = getTemplate()
    eng_chosen, eng_num = getEngagement()
    print(subjects[temp_num])
    print(bodytexts[temp_num])
    print(manager[eng_num])
    print(reviewer[eng_num])
    print(preparer[eng_num])

# create window
root = Tk()
root.title("E-mail generator")
root.geometry("700x800")
root.config(background="skyblue")


# create notebook to make separate tabs
notebook = ttk.Notebook(root, width=600)
notebook.place(relx=0.5, rely=0.01, anchor=N)

# create the 3 different tabs and the frames that each will contain
# TAB 1 for the templates

template_frame = Frame(root, relief=RIDGE, borderwidth=5, background="Pink")
template_frame.place()

first_frame = Frame(template_frame, relief=RIDGE, borderwidth=5, background="green")
first_frame.pack(fill=X)

frame1 = Frame(first_frame)
frame1.pack(fill=X, side=LEFT)

frame2 = Frame(first_frame, relief=RIDGE, borderwidth=5)
frame2.pack(fill=BOTH, side=RIGHT)

frame3 = Frame(template_frame, relief=RIDGE, borderwidth=5)
frame3.pack()
frame4 = Frame(template_frame, relief=RIDGE, borderwidth=5)
frame4.pack()
label = Label(frame1, text="Hello!")
label.pack()

# TAB 2 for the residency
residency_frame = Frame(root, relief=RIDGE, borderwidth=5)
residency_frame.pack()
label = Label(residency_frame, text="Üdvözöllek a reziség egyeztető e-mail generáló programban!")
label.pack()

# # TAB 3 for the advisory
advisory_frame = Frame(root, relief=RIDGE, borderwidth=5)
advisory_frame.pack()
label = Label(advisory_frame, text="Tanácsadás generáló programban!")
label.pack()

# add tabs to the notepad
notebook.add(template_frame, text="Template")
notebook.add(residency_frame, text="Residency")
notebook.add(advisory_frame, text="Advisory")

# TEMPLATE
# fill up the template tab

template_label = Label(frame1, text="Template?")
template_label.pack()
template_list = ttk.Combobox(frame1, values = templates)
template_list.pack()


engagement_label = Label(frame2, text="Engagement?")
engagement_label.pack()
engagement_list = ttk.Combobox(frame2, values = eng_name)
engagement_list.pack()


generate_button = Button(frame3, text="Generálás!", command=lambda: [getTemplate(), getEngagement()])
generate_button.pack()


another_button = Button(frame4, text="Print!", command=getRemaining)
another_button.pack()

# open up new e-mail
# outlook = win32com.client.Dispatch('outlook.application')
# mail = outlook.CreateItem(0)

# mail.To = 'adamnorbert90@gmail.com'
# mail.Subject = 'Próba'
# mail.Body = 'Próba'

# mail.Display(True)

root.mainloop()